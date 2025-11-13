"""Enhanced Excel operations module for reading, sorting, and saving workbooks."""
import os
import subprocess
from openpyxl import load_workbook
from sheet_rules import apply_template
from backup_util import make_backup

class ExcelHandler:
    """Handles Excel workbook operations safely and with debug logging."""
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None

    def load_workbook(self) -> bool:
        """Loads the Excel workbook.
        Before calling openpyxl.load_workbook, attempt to open the file in binary
        read+write mode. On Windows, if Excel has the file open, this will raise
        a PermissionError. We use that to detect "file is open/locked" and set
        `self.file_open_locked` so the UI can show a helpful warning."""
        # reset flag each time
        self.file_open_locked = False

        if not os.path.exists(self.file_path):
            print(f"[ERROR] File not found: {self.file_path}")
            return False

        # Quick check: try to open file in binary read+write mode. If file is locked
        # by Excel on Windows, this usually raises PermissionError.
        try:
            # 'r+b' opens for reading and writing in binary; it will fail if file is locked
            with open(self.file_path, "r+b"):
                pass
        except PermissionError:
            # File is locked by another program (often Excel). Set flag and return False.
            self.file_open_locked = True
            print(f"[ERROR] Permission denied (file likely open in Excel): {self.file_path}")
            return False
        except OSError as err:
            # Could not open file for other OS-related reasons; still try load_workbook below
            print(f"[WARNING] Could not perform exclusive open check: {err}. Will attempt to load anyway.")

        # If the quick-check passed (or only raised non-blocking OSError), try loading
        try:
            # read_only=False ensures the workbook is editable by openpyxl
            self.workbook = load_workbook(filename=self.file_path, read_only=False, data_only=False)
            print(f"[INFO] Workbook loaded successfully: {self.file_path}")
            return True
        except PermissionError:
            # An unexpected permission error while loading — treat as locked file
            self.file_open_locked = True
            print(f"[ERROR] Permission denied when loading workbook (file may be open): {self.file_path}")
            return False
        except Exception as err:
            print(f"[ERROR while loading workbook] {err}")
            return False

    def get_sheet_names(self) -> list:
        """Returns a list of sheet names."""
        if self.workbook:
            return self.workbook.sheetnames
        return []

    def sort_sheets_alphabetically(self) -> bool:
        """Sorts visible sheets alphabetically (A–Z) while keeping hidden ones in place."""
        if not self.workbook:
            print("[ERROR] Workbook not loaded before sorting.")
            return False

        try:
            # Separate visible and hidden sheets
            visible_sheets = [
                ws for ws in self.workbook._sheets if ws.sheet_state == "visible"
            ]
            hidden_sheets = [
                ws for ws in self.workbook._sheets if ws.sheet_state != "visible"
            ]

            # Sort visible ones alphabetically
            visible_sheets.sort(key=lambda ws: ws.title.lower())

            # Recombine sheets: visible first, hidden last (to preserve state)
            self.workbook._sheets = visible_sheets + hidden_sheets

            print(
                "[INFO] Sheets sorted alphabetically: "
                f"{[ws.title for ws in visible_sheets]}"
            )
            return True
            #subprocess.Popen(["start", path], shell=True)
        except Exception as err:
            print(f"[ERROR while sorting sheets] {err}")
            return False

    def apply_custom_sort(self, key_func) -> bool:
        """Sort visible sheets using provided key function."""
        if not self.workbook:
            print("[ERROR] Workbook not loaded before custom sort.")
            return False
        try:
            visible = [ws for ws in self.workbook._sheets if ws.sheet_state == "visible"]
            hidden = [ws for ws in self.workbook._sheets if ws.sheet_state != "visible"]
            visible.sort(key=key_func)
            self.workbook._sheets = visible + hidden
            return True
        except Exception as err:
            print(f"[ERROR while custom sorting] {err}")
            return False

    def rename_sheets_with_template(self, template: str) -> bool:
        """Rename sheets using a template string safely."""
        if not self.workbook:
            print("[ERROR] Workbook not loaded before renaming.")
            return False
        try:
            for i, ws in enumerate(self.workbook._sheets, start=1):
                new_name = apply_template(ws.title, template, i)
                # openpyxl will raise if invalid name; handle gracefully
                ws.title = new_name
            return True
        except Exception as err:
            print(f"[ERROR while renaming sheets] {err}")
            return False

    def backup_before_save(self) -> str:
        """Create a backup and return path or empty string."""
        try:
            return make_backup(self.file_path)
        except Exception:
            return ""
    # --- end insertion

    def save_workbook(self) -> bool:
        """Saves the workbook to the same file path with permission handling."""
        if not self.workbook:
            print("[ERROR] Workbook not loaded before saving.")
            return False

        try:
            # Check write permission on the file (folder)
            if not os.access(self.file_path, os.W_OK):
                print(f"[ERROR] No write permission for: {self.file_path}")
                return False

            # Attempt to save
            self.workbook.save(self.file_path)
            print(f"[INFO] Workbook saved successfully: {self.file_path}")
            return True

            backup = self.backup_before_save()
            if backup:
                print(f"[INFO] Backup created: {backup}")
            self.workbook.save(self.file_path)

        except PermissionError:
            print(
                "[ERROR] Cannot save — file is open in another program (e.g., Excel). "
                "Close it and retry."
            )
            return False

        except Exception as err:
            print(f"[ERROR while saving workbook] {err}")
            return False

    def save_as(self, new_path: str) -> bool:
        """Saves workbook as a new file."""
        try:
            self.workbook.save(new_path)
            print(f"[INFO] Workbook saved as: {new_path}")
            return True
        except Exception as err:
            print(f"[ERROR] Save-As failed: {err}")
            return False
