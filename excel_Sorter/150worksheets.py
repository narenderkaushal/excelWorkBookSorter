""" from openpyxl import Workbook

wb = Workbook()
# Remove the default sheet
wb.remove(wb.active)

# Add 150 blank sheets
for i in range(1, 151):
    wb.create_sheet(title=f"Sheet{i}")

wb.save("workbook_150_sheets.xlsx") """

from openpyxl import Workbook

wb = Workbook()
wb.remove(wb.active)

for i in range(1, 45):
    sheet = wb.create_sheet(title=f"TestData_{i}")
    # Sample table headers
    headers = ["ID", "Name", "Value"]
    sheet.append(headers)
    # Add 35 rows of sample data
    for row in range(1, 36):
        sheet.append([row, f"Item{row}", row * i])

wb.save("workbook_44_sheets_with_tables.xlsx")